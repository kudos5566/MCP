#!/usr/bin/env python3
"""
Excel报告生成器模块

提供安全扫描结果的Excel报告生成功能，包括：
- 扫描摘要工作表
- 详细扫描结果工作表
- 漏洞详情工作表
- 二次扫描结果工作表
- 字符清理和格式化功能
"""

import datetime
import re
from io import BytesIO
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelReportGenerator:
    """生成专业的安全扫描Excel报告"""
    
    def __init__(self):
        self.wb = Workbook()
        self.setup_styles()
    
    def setup_styles(self):
        """设置Excel样式"""
        # 标题样式
        self.title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        self.title_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        
        # 表头样式
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # 数据样式
        self.data_font = Font(name='Arial', size=10)
        self.data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # 边框样式
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 风险等级颜色
        self.risk_colors = {
            'High': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
            'Medium': PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid'),
            'Low': PatternFill(start_color='6BCF7F', end_color='6BCF7F', fill_type='solid'),
            'Info': PatternFill(start_color='74C0FC', end_color='74C0FC', fill_type='solid')
        }
    
    def _clean_text_for_excel(self, text: str) -> str:
        """清理文本中Excel不支持的字符"""
        if not text:
            return text
        
        # 移除控制字符和非打印字符，但保留常见的空白字符
        # 保留制表符(\t)、换行符(\n)、回车符(\r)和普通空格
        cleaned = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]', '', text)
        
        # 移除或替换ASCII艺术字符和特殊符号
        # 替换常见的ASCII艺术字符
        ascii_art_chars = ['│', '┌', '┐', '└', '┘', '├', '┤', '┬', '┴', '┼', '─', '║', '╔', '╗', '╚', '╝', '╠', '╣', '╦', '╩', '╬', '═']
        for char in ascii_art_chars:
            cleaned = cleaned.replace(char, '|')
        
        # 移除其他可能有问题的Unicode字符，保留中文字符
        cleaned = re.sub(r'[^\x20-\x7E\t\n\r\u4e00-\u9fff]', '', cleaned)
        
        return cleaned
    
    def create_summary_sheet(self, scan_results: Dict[str, Any]):
        """创建扫描摘要工作表"""
        ws = self.wb.active
        ws.title = "扫描摘要"
        
        # 报告标题
        ws['A1'] = "安全扫描报告"
        ws['A1'].font = self.title_font
        ws['A1'].fill = self.title_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        
        # 基本信息
        row = 3
        ws[f'A{row}'] = "扫描目标:"
        ws[f'B{row}'] = scan_results.get('target', 'N/A')
        row += 1
        ws[f'A{row}'] = "扫描时间:"
        ws[f'B{row}'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        ws[f'A{row}'] = "扫描工具:"
        tools_used = [step['tool'] for step in scan_results.get('scan_sequence', [])]
        ws[f'B{row}'] = ', '.join(tools_used)
        
        # 扫描统计
        row += 2
        ws[f'A{row}'] = "扫描统计"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        stats = self._calculate_scan_stats(scan_results)
        for key, value in stats.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
    
    def create_detailed_results_sheet(self, scan_results: Dict[str, Any]):
        """创建详细结果工作表"""
        ws = self.wb.create_sheet("详细扫描结果")
        
        # 表头
        headers = ['工具', '命令', '执行状态', '输出摘要', '错误信息', '执行时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # 数据行
        row = 2
        for step in scan_results.get('scan_sequence', []):
            ws.cell(row=row, column=1, value=step.get('tool', ''))
            ws.cell(row=row, column=2, value=self._clean_text_for_excel(step.get('command', '')))
            
            result = step.get('result', {})
            status = "成功" if result.get('success', False) else "失败"
            ws.cell(row=row, column=3, value=status)
            
            # 输出摘要（截取前200字符并清理特殊字符）
            stdout = result.get('stdout', '')
            summary = (stdout[:200] + '...') if len(stdout) > 200 else stdout
            cleaned_summary = self._clean_text_for_excel(summary)
            ws.cell(row=row, column=4, value=cleaned_summary)
            
            stderr = self._clean_text_for_excel(result.get('stderr', ''))
            ws.cell(row=row, column=5, value=stderr)
            
            summary_text = self._clean_text_for_excel(step.get('summary', ''))
            ws.cell(row=row, column=6, value=summary_text)
            
            # 应用样式
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.font = self.data_font
                cell.border = self.border
                cell.alignment = self.data_alignment
            
            row += 1
        
        # 设置列宽
        column_widths = [15, 50, 10, 60, 30, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
    
    def create_vulnerabilities_sheet(self, scan_results: Dict[str, Any]):
        """创建漏洞详情工作表"""
        ws = self.wb.create_sheet("发现的漏洞")
        
        # 表头
        headers = ['漏洞类型', '风险等级', '影响组件', '描述', '修复建议', '发现工具']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # 提取漏洞信息
        vulnerabilities = self._extract_vulnerabilities(scan_results)
        
        row = 2
        for vuln in vulnerabilities:
            ws.cell(row=row, column=1, value=self._clean_text_for_excel(vuln.get('type', '')))
            ws.cell(row=row, column=2, value=vuln.get('risk_level', ''))
            ws.cell(row=row, column=3, value=self._clean_text_for_excel(vuln.get('component', '')))
            ws.cell(row=row, column=4, value=self._clean_text_for_excel(vuln.get('description', '')))
            ws.cell(row=row, column=5, value=self._clean_text_for_excel(vuln.get('remediation', '')))
            ws.cell(row=row, column=6, value=vuln.get('tool', ''))
            
            # 应用样式和风险等级颜色
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.font = self.data_font
                cell.border = self.border
                cell.alignment = self.data_alignment
                
                # 风险等级着色
                if col == 2:  # 风险等级列
                    risk_level = vuln.get('risk_level', '')
                    if risk_level in self.risk_colors:
                        cell.fill = self.risk_colors[risk_level]
            
            row += 1
        
        # 设置列宽
        column_widths = [20, 12, 25, 50, 50, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
    
    def create_secondary_scans_sheet(self, scan_results: Dict[str, Any]):
        """创建二次扫描结果工作表"""
        secondary_scans = scan_results.get('secondary_scans', [])
        if not secondary_scans:
            return
        
        ws = self.wb.create_sheet("二次扫描结果")
        
        # 表头
        headers = ['目标URL', '扫描工具', '发现问题', '风险等级', '详细信息']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        for scan in secondary_scans:
            ws.cell(row=row, column=1, value=scan.get('target_url', ''))
            ws.cell(row=row, column=2, value=scan.get('tool', ''))
            ws.cell(row=row, column=3, value=scan.get('issues_found', 0))
            ws.cell(row=row, column=4, value=scan.get('max_risk_level', 'Info'))
            
            # 详细信息摘要
            result = scan.get('result', {})
            details = result.get('stdout', '')[:300] + '...' if len(result.get('stdout', '')) > 300 else result.get('stdout', '')
            cleaned_details = self._clean_text_for_excel(details)
            ws.cell(row=row, column=5, value=cleaned_details)
            
            # 应用样式
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.font = self.data_font
                cell.border = self.border
                cell.alignment = self.data_alignment
            
            row += 1
        
        # 设置列宽
        column_widths = [40, 15, 15, 15, 60]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
    
    def _calculate_scan_stats(self, scan_results: Dict[str, Any]) -> Dict[str, Any]:
        """计算扫描统计信息"""
        stats = {
            "执行的扫描步骤": len(scan_results.get('scan_sequence', [])),
            "成功的扫描": len([s for s in scan_results.get('scan_sequence', []) if s.get('result', {}).get('success', False)]),
            "失败的扫描": len([s for s in scan_results.get('scan_sequence', []) if not s.get('result', {}).get('success', False)]),
            "二次扫描数量": len(scan_results.get('secondary_scans', [])),
            "发现的漏洞": len(self._extract_vulnerabilities(scan_results)),
            "错误数量": len(scan_results.get('errors', []))
        }
        return stats
    
    def _extract_vulnerabilities(self, scan_results: Dict[str, Any]) -> List[Dict[str, Any]]:
        """从扫描结果中提取漏洞信息和发现项"""
        vulnerabilities = []
        
        # 从各个工具的输出中提取漏洞
        for step in scan_results.get('scan_sequence', []):
            tool = step.get('tool', '')
            result = step.get('result', {})
            stdout = result.get('stdout', '')
            
            if tool == 'nikto' and stdout:
                # 解析Nikto输出
                nikto_vulns = self._parse_nikto_vulnerabilities(stdout)
                vulnerabilities.extend(nikto_vulns)
            elif tool == 'nmap' and stdout:
                # 解析Nmap输出中的CVE和开放端口
                nmap_vulns = self._parse_nmap_vulnerabilities(stdout)
                vulnerabilities.extend(nmap_vulns)
                # 解析开放端口
                port_findings = self._parse_nmap_ports(stdout)
                vulnerabilities.extend(port_findings)
            elif tool == 'gobuster' and stdout:
                # 解析Gobuster发现的路径
                gobuster_findings = self._parse_gobuster_paths(stdout)
                vulnerabilities.extend(gobuster_findings)
            elif tool == 'dirb' and stdout:
                # 解析Dirb发现的路径
                dirb_findings = self._parse_dirb_paths(stdout)
                vulnerabilities.extend(dirb_findings)
            elif tool == 'urlfinder' and stdout:
                # 解析URLFinder发现的URL
                urlfinder_findings = self._parse_urlfinder_urls(stdout)
                vulnerabilities.extend(urlfinder_findings)
            elif tool == 'sqlmap' and stdout:
                # 解析SQLMap发现的注入点
                sqlmap_findings = self._parse_sqlmap_injections(stdout)
                vulnerabilities.extend(sqlmap_findings)
            elif tool == 'wpscan' and stdout:
                # 解析WPScan发现的WordPress漏洞
                wpscan_findings = self._parse_wpscan_vulnerabilities(stdout)
                vulnerabilities.extend(wpscan_findings)
            elif tool == 'hydra' and stdout:
                # 解析Hydra破解结果
                hydra_findings = self._parse_hydra_credentials(stdout)
                vulnerabilities.extend(hydra_findings)
            elif tool == 'john' and stdout:
                # 解析John破解结果
                john_findings = self._parse_john_credentials(stdout)
                vulnerabilities.extend(john_findings)
        
        # 从二次扫描中提取漏洞
        for scan in scan_results.get('secondary_scans', []):
            result = scan.get('result', {})
            stdout = result.get('stdout', '')
            if stdout:
                secondary_vulns = self._parse_nikto_vulnerabilities(stdout)
                for vuln in secondary_vulns:
                    vuln['component'] = scan.get('target_url', '')
                vulnerabilities.extend(secondary_vulns)
        
        return vulnerabilities
    
    def _parse_nikto_vulnerabilities(self, nikto_output: str) -> List[Dict[str, Any]]:
        """解析Nikto输出中的漏洞"""
        vulnerabilities = []
        lines = nikto_output.split('\n')
        
        for line in lines:
            if '+ ' in line and any(keyword in line.lower() for keyword in ['vulnerability', 'security', 'risk', 'exploit', 'cve']):
                vuln = {
                    'type': 'Web漏洞',
                    'risk_level': self._assess_risk_level(line),
                    'component': 'Web服务',
                    'description': line.strip(),
                    'remediation': '请根据具体漏洞类型进行修复',
                    'tool': 'nikto'
                }
                vulnerabilities.append(vuln)
        
        return vulnerabilities
    
    def _parse_nmap_vulnerabilities(self, nmap_output: str) -> List[Dict[str, Any]]:
        """解析Nmap输出中的CVE漏洞"""
        vulnerabilities = []
        lines = nmap_output.split('\n')
        
        for line in lines:
            if 'CVE-' in line:
                cve_match = re.search(r'CVE-\d{4}-\d{4,}', line)
                if cve_match:
                    cve_id = cve_match.group()
                    vuln = {
                        'type': 'CVE漏洞',
                        'risk_level': 'Medium',
                        'component': '网络服务',
                        'description': f'发现CVE漏洞: {cve_id}',
                        'remediation': f'请查看CVE详情并应用相应补丁: {cve_id}',
                        'tool': 'nmap'
                    }
                    vulnerabilities.append(vuln)
        
        return vulnerabilities
    
    def _assess_risk_level(self, description: str) -> str:
        """评估风险等级"""
        description_lower = description.lower()
        
        high_risk_keywords = ['critical', 'high', 'exploit', 'rce', 'sql injection', 'xss']
        medium_risk_keywords = ['medium', 'warning', 'disclosure', 'bypass']
        
        if any(keyword in description_lower for keyword in high_risk_keywords):
            return 'High'
        elif any(keyword in description_lower for keyword in medium_risk_keywords):
            return 'Medium'
        else:
            return 'Low'
    
    def _parse_nmap_ports(self, nmap_output: str) -> List[Dict[str, Any]]:
        """解析Nmap发现的开放端口"""
        findings = []
        lines = nmap_output.split('\n')
        
        for line in lines:
            if '/tcp' in line and 'open' in line:
                parts = line.split()
                if len(parts) >= 3:
                    port = parts[0].split('/')[0]
                    service = parts[2] if len(parts) > 2 else 'unknown'
                    
                    finding = {
                        'type': '开放端口',
                        'risk_level': 'Info',
                        'component': f'端口 {port}',
                        'description': f'发现开放端口: {port} ({service})',
                        'remediation': '检查是否需要关闭不必要的端口',
                        'tool': 'nmap'
                    }
                    findings.append(finding)
        
        return findings
    
    def _parse_gobuster_paths(self, gobuster_output: str) -> List[Dict[str, Any]]:
        """解析Gobuster发现的路径"""
        findings = []
        lines = gobuster_output.split('\n')
        
        for line in lines:
            if line.startswith('/'):
                parts = line.split()
                if len(parts) >= 2:
                    path = parts[0]
                    status = parts[1] if len(parts) > 1 else 'unknown'
                    
                    finding = {
                        'type': '发现路径',
                        'risk_level': 'Info',
                        'component': path,
                        'description': f'发现可访问路径: {path} (状态码: {status})',
                        'remediation': '检查路径是否应该公开访问',
                        'tool': 'gobuster'
                    }
                    findings.append(finding)
        
        return findings
    
    def _parse_dirb_paths(self, dirb_output: str) -> List[Dict[str, Any]]:
        """解析Dirb发现的路径"""
        findings = []
        lines = dirb_output.split('\n')
        
        for line in lines:
            if '==>' in line and 'http' in line:
                url_match = re.search(r'http[s]?://[^\s]+', line)
                if url_match:
                    url = url_match.group()
                    
                    finding = {
                        'type': '发现路径',
                        'risk_level': 'Info',
                        'component': url,
                        'description': f'发现可访问路径: {url}',
                        'remediation': '检查路径是否应该公开访问',
                        'tool': 'dirb'
                    }
                    findings.append(finding)
        
        return findings
    
    def _parse_urlfinder_urls(self, urlfinder_output: str) -> List[Dict[str, Any]]:
        """解析URLFinder发现的URL"""
        findings = []
        lines = urlfinder_output.split('\n')
        
        for line in lines:
            if line.startswith('http'):
                url = line.strip()
                
                finding = {
                    'type': '发现URL',
                    'risk_level': 'Info',
                    'component': url,
                    'description': f'发现URL: {url}',
                    'remediation': '检查URL是否包含敏感信息',
                    'tool': 'urlfinder'
                }
                findings.append(finding)
        
        return findings
    
    def _parse_sqlmap_injections(self, sqlmap_output: str) -> List[Dict[str, Any]]:
        """解析SQLMap发现的注入点"""
        findings = []
        lines = sqlmap_output.split('\n')
        
        for line in lines:
            if 'vulnerable' in line.lower() or 'injection' in line.lower():
                finding = {
                    'type': 'SQL注入',
                    'risk_level': 'High',
                    'component': 'Web应用',
                    'description': f'发现SQL注入漏洞: {line.strip()}',
                    'remediation': '使用参数化查询，验证输入数据',
                    'tool': 'sqlmap'
                }
                findings.append(finding)
        
        return findings
    
    def _parse_wpscan_vulnerabilities(self, wpscan_output: str) -> List[Dict[str, Any]]:
        """解析WPScan发现的WordPress漏洞"""
        findings = []
        lines = wpscan_output.split('\n')
        
        for line in lines:
            if 'vulnerability' in line.lower() or 'cve' in line.lower():
                finding = {
                    'type': 'WordPress漏洞',
                    'risk_level': self._assess_risk_level(line),
                    'component': 'WordPress',
                    'description': line.strip(),
                    'remediation': '更新WordPress核心、主题和插件',
                    'tool': 'wpscan'
                }
                findings.append(finding)
        
        return findings
    
    def _parse_hydra_credentials(self, hydra_output: str) -> List[Dict[str, Any]]:
        """解析Hydra破解的凭据"""
        findings = []
        lines = hydra_output.split('\n')
        
        for line in lines:
            if 'login:' in line and 'password:' in line:
                finding = {
                    'type': '弱凭据',
                    'risk_level': 'High',
                    'component': '认证系统',
                    'description': f'发现弱凭据: {line.strip()}',
                    'remediation': '使用强密码策略，启用多因素认证',
                    'tool': 'hydra'
                }
                findings.append(finding)
        
        return findings
    
    def _parse_john_credentials(self, john_output: str) -> List[Dict[str, Any]]:
        """解析John破解的密码"""
        findings = []
        lines = john_output.split('\n')
        
        for line in lines:
            if ':' in line and len(line.split(':')) >= 2:
                finding = {
                    'type': '弱密码',
                    'risk_level': 'High',
                    'component': '密码系统',
                    'description': f'发现弱密码: {line.strip()}',
                    'remediation': '使用强密码策略，定期更换密码',
                    'tool': 'john'
                }
                findings.append(finding)
        
        return findings
    
    def generate_report(self, scan_results: Dict[str, Any]) -> BytesIO:
        """生成完整的Excel报告"""
        # 创建各个工作表
        self.create_summary_sheet(scan_results)
        self.create_detailed_results_sheet(scan_results)
        self.create_vulnerabilities_sheet(scan_results)
        self.create_secondary_scans_sheet(scan_results)
        
        # 保存到内存
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        return output